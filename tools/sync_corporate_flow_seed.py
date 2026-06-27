from __future__ import annotations

import hashlib
import json
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(r"D:\Projects\SpeedEmulator")
PERSONAL_SOURCE_DIR = Path(r"D:\Projects\财务系统开发资料\参照明细")
SOURCE_DIR = Path(r"D:\Projects\财务系统开发资料\参照明细\对公银行")
COLUMN_OUTPUT = PROJECT_DIR / "Data" / "zhencheng-flow-rule-columns.json"
SEED_OUTPUT = PROJECT_DIR / "Data" / "zhencheng-flow-generation-seed.json"

REFERENCE_KIND = "Reference"
CONST_KIND = "Const"

CORPORATE_BANK_IDS = {
    "对公农商": 101,
    "民生对公": 102,
    "中信对公": 103,
    "工行对公": 104,
    "农行对公": 105,
    "中行对公": 106,
    "光大对公": 107,
    "平安对公": 108,
    "广发对公": 109,
    "浦发对公": 110,
    "华夏对公": 111,
    "兴业对公": 112,
    "建行对公": 113,
    "邮政对公": 114,
    "交行对公": 115,
    "招行对公": 116,
}

PERSONAL_BANK_SEED_KEYS = {
    "支付宝": "1",
    "微信": "2",
    "农行": "3",
    "工行": "4",
    "光大": "5",
    "广发": "6",
    "华夏": "7",
    "建行": "8",
    "交行": "9",
    "民生": "10",
    "平安": "11",
    "浦发": "12",
    "兴业": "13",
    "邮政": "14",
    "个人农商": "15",
    "农商": "15",
    "招行": "16",
    "中信": "17",
    "中行": "18",
}

PERSONAL_BANK_IDS = {
    name: int(bank_id)
    for name, bank_id in PERSONAL_BANK_SEED_KEYS.items()
    if name != "农商"
}

NAME_ALIASES = {
    "农商对公": "对公农商",
}

BOOL_FIELDS = {"isCheck", "tradeHoliday", "tradeWeekend"}
MONEY_FIELDS = {
    "minMoney",
    "maxMoney",
    "crossBankRate",
    "crossBankMin",
    "crossBankMax",
    "offSiteBankRate",
    "offSiteBankMin",
    "offSiteBankMax",
    "creditAmount",
    "debitAmount",
    "balanceAmount",
}
INT_FIELDS = {"floutLength", "startDay", "endDay", "percentMonth"}


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def trim(values: list[Any]) -> list[Any]:
    while values and values[-1] in ("", None):
        values.pop()
    return values


def find_header_row(sheet: Any) -> tuple[int, list[str]]:
    best: tuple[tuple[int, int], int, list[Any]] | None = None

    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        values = trim([normalize_value(cell.value) for cell in row])
        text_count = sum(1 for value in values if isinstance(value, str) and value)
        non_empty = sum(1 for value in values if value not in ("", None))

        if text_count < 2 or non_empty < 2:
            continue

        score = (text_count, non_empty)
        if best is None or score > best[0]:
            best = (score, row_index, values)

        if text_count >= 4:
            return row_index, [str(value) for value in values]

    if best is None:
        return 1, []

    return best[1], [str(value) for value in best[2]]


def classify(path: Path) -> tuple[str, str]:
    stem = path.stem.strip()
    kind = CONST_KIND if "固定" in stem else REFERENCE_KIND
    bank_name = stem.replace("固定日期增加项目", "")
    bank_name = bank_name.replace("固定日期增加项", "")
    bank_name = bank_name.replace("-固定", "")
    bank_name = bank_name.replace("、固定", "")
    bank_name = bank_name.replace("固定", "")
    bank_name = bank_name.strip(" -_、")
    bank_name = NAME_ALIASES.get(bank_name, bank_name)
    return bank_name, kind


def read_workbook(path: Path) -> dict[str, Any]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, data_only=True, read_only=True)

    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_row, headers = find_header_row(sheet)
        rows = []

        for row in sheet.iter_rows(min_row=header_row + 1):
            values = trim([normalize_value(cell.value) for cell in row])
            if any(value not in ("", None) for value in values):
                rows.append(values)

        bank_name, kind = classify(path)
        return {
            "file": str(path),
            "bank": bank_name,
            "kind": kind,
            "headers": ensure_id_header(headers),
            "sourceHeaders": headers,
            "rows": rows,
        }
    finally:
        workbook.close()


def ensure_id_header(headers: list[str]) -> list[str]:
    cleaned = [str(header).strip() for header in headers if str(header).strip()]
    if not cleaned or cleaned[0] != "ID":
        return ["ID", *cleaned]
    return cleaned


def extra_field(bank_name: str, kind: str, column_name: str, column_index: int) -> str:
    raw = f"{bank_name}|{kind}|{column_index}|{column_name}"
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest().upper()[:12]
    return f"RuleField_{digest}"


def resolve_field(column_name: str, kind: str) -> tuple[str | None, str]:
    if kind == REFERENCE_KIND and column_name == "每月出现次数":
        return "percentMonth", "Text"

    if kind == CONST_KIND:
        if column_name == "固定添加日":
            return "fixDay", "Text"
        if column_name == "次数":
            return "reCnt", "Text"

    mapping = {
        "ID": ("id", "Text"),
        "选择": ("isCheck", "Boolean"),
        "收支属性": ("incomeAttribute", "Text"),
        "最小金额": ("minMoney", "Money"),
        "最小金颔": ("minMoney", "Money"),
        "最小金颜": ("minMoney", "Money"),
        "最大金额": ("maxMoney", "Money"),
        "最大金颜": ("maxMoney", "Money"),
        "小数位": ("floutLength", "Text"),
        "开始时间": ("startDay", "Text"),
        "结束时间": ("endDay", "Text"),
        "节假日交易": ("tradeHoliday", "Boolean"),
        "周六日交易": ("tradeWeekend", "Boolean"),
        "备注": ("remark", "Text"),
        "附言": ("remark", "Text"),
        "留言": ("remark", "Text"),
        "转账附言": ("remark", "Text"),
        "流水号": ("serialNum", "Text"),
        "交易流水": ("serialNum", "Text"),
        "交易流水号": ("serialNum", "Text"),
        "核心流水号": ("serialNum", "Text"),
        "资金渠道": ("tradeChannel", "Text"),
        "交易渠道": ("tradeChannel", "Text"),
        "渠道": ("tradeChannel", "Text"),
        "交易渠道英文": ("tradeChannelEn", "Text"),
        "交易对方": ("oppositeUsername", "Text"),
        "对方户名": ("oppositeUsername", "Text"),
        "对方名称": ("oppositeUsername", "Text"),
        "对手户名": ("oppositeUsername", "Text"),
        "对手方户名": ("oppositeUsername", "Text"),
        "对方姓名": ("oppositeUsername", "Text"),
        "对方账户名": ("oppositeUsername", "Text"),
        "交易对手名称": ("oppositeUsername", "Text"),
        "对方账号名称": ("oppositeUsername", "Text"),
        "户名": ("oppositeUsername", "Text"),
        "对方账号": ("oppositeAccount", "Text"),
        "对手账号": ("oppositeAccount", "Text"),
        "对手方账号": ("oppositeAccount", "Text"),
        "交易对手账号": ("oppositeAccount", "Text"),
        "对方卡号账号": ("oppositeAccount", "Text"),
        "对方开户行": ("oppositeBank", "Text"),
        "对方银行": ("oppositeBank", "Text"),
        "对手银行": ("oppositeBank", "Text"),
        "对方行名": ("oppositeBank", "Text"),
        "对方行号": ("oppositeBank", "Text"),
        "对方开户行银联": ("oppositeBank", "Text"),
        "商家订单号": ("merchantName", "Text"),
        "商户单号": ("merchantName", "Text"),
        "商户名称": ("merchantName", "Text"),
        "支付宝分类": ("incomeType", "Text"),
        "交易分类": ("incomeType", "Text"),
        "收支其他": ("incomeType", "Text"),
        "账号": ("account", "Text"),
        "卡号": ("account", "Text"),
        "客户账号": ("account", "Text"),
        "交易账号": ("account", "Text"),
        "账户代码": ("account", "Text"),
        "应用号": ("appNum", "Text"),
        "序号": ("sequenceNum", "Text"),
        "币种": ("currency", "Text"),
        "货币": ("currency", "Text"),
        "贷币": ("currency", "Text"),
        "交易币种": ("tradeCurrency", "Text"),
        "钞汇": ("cashCheck", "Text"),
        "交易方式": ("cashCheck", "Text"),
        "交易代码": ("tradeCode", "Text"),
        "交易码": ("tradeCode", "Text"),
        "注释": ("tradeExplain", "Text"),
        "交易说明": ("tradeExplain", "Text"),
        "交易描述": ("tradeExplain", "Text"),
        "存期": ("depositTerm", "Text"),
        "约转期": ("agreedTerm", "Text"),
        "通知种类发行代": ("noticeType", "Text"),
        "地区号": ("areaNum", "Text"),
        "网点号": ("netNum", "Text"),
        "机构号": ("netNum", "Text"),
        "机构码": ("netNum", "Text"),
        "交易机构号": ("netNum", "Text"),
        "操作员": ("operator", "Text"),
        "柜员": ("operator", "Text"),
        "业务柜员": ("operator", "Text"),
        "交易柜员": ("operator", "Text"),
        "柜员号": ("operatorNum", "Text"),
        "柜员流水": ("operatorNum", "Text"),
        "柜员交易号": ("operatorNum", "Text"),
        "交易柜员号": ("operatorNum", "Text"),
        "机构柜员流水": ("operatorNum", "Text"),
        "界面": ("interfacePage", "Text"),
        "交易场所": ("tradePlace", "Text"),
        "交易地点": ("tradePlace", "Text"),
        "地点": ("tradePlace", "Text"),
        "交易网点": ("tradePlace", "Text"),
        "交易行所": ("tradePlace", "Text"),
        "商户网点号及名称": ("tradePlace", "Text"),
        "商户网点号及名": ("tradePlace", "Text"),
        "交易机构名称": ("tradePlace", "Text"),
        "交易地点/附言": ("tradePlace", "Text"),
        "摘要": ("productBrief", "Text"),
        "交易摘要": ("productBrief", "Text"),
        "产品摘要": ("productBrief", "Text"),
        "银行摘要": ("productBrief", "Text"),
        "账单摘要": ("productBrief", "Text"),
        "摘要代码": ("productBrief", "Text"),
        "记账信息": ("productBrief", "Text"),
        "产品名称": ("productName", "Text"),
        "交易名称": ("productName", "Text"),
        "交易种类": ("productName", "Text"),
        "业务类型": ("productName", "Text"),
        "交易类型": ("productName", "Text"),
        "业务产品种类": ("productType", "Text"),
        "产品业务种类": ("productType", "Text"),
        "用途": ("usage", "Text"),
        "交易用途": ("usage", "Text"),
        "分户序号": ("subAccountNum", "Text"),
        "子账户序号": ("subAccountNum", "Text"),
        "账户序号": ("accountNum", "Text"),
        "账号序号": ("accountNum", "Text"),
        "凭证类型": ("voucherType", "Text"),
        "凭证种类": ("voucherType", "Text"),
        "凭证号": ("voucherNum", "Text"),
        "凭证号码": ("voucherNum", "Text"),
        "凭证": ("voucherNum", "Text"),
        "票据号": ("voucherNum", "Text"),
        "传票号": ("voucherNum", "Text"),
        "凭证序号": ("voucherNum", "Text"),
        "凭证号码业务": ("voucherNum", "Text"),
        "外部系统流水": ("voucherNum", "Text"),
        "日志号": ("logNum", "Text"),
        "年份": ("year", "Text"),
        "借方发生额": ("debitAmount", "Money"),
        "贷方发生额": ("creditAmount", "Money"),
        "余额": ("balanceAmount", "Money"),
        "回单编号": ("receiptNum", "Text"),
        "全局路由号": ("receiptNum", "Text"),
    }

    return mapping.get(column_name, (None, "Text"))


def should_use_extra(field: str, used_fixed_fields: set[str]) -> bool:
    if field == "id":
        return False
    if field in used_fixed_fields:
        return True

    used_fixed_fields.add(field)
    return False


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0

    text = str(value).strip().lower()
    return text in {"true", "1", "是", "y", "yes", "选中"}


def to_number(value: Any) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def to_int(value: Any) -> int | None:
    number = to_number(value)
    if number is None:
        return None
    return int(number)


def to_text(value: Any) -> str:
    if value in ("", None):
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def convert_value(field: str, value: Any) -> Any:
    if field in BOOL_FIELDS:
        return to_bool(value)
    if field in MONEY_FIELDS:
        return to_number(value)
    if field in INT_FIELDS:
        return to_int(value)
    return to_text(value)


def create_rule(
    bank_name: str,
    bank_id: int,
    kind: str,
    headers: list[str],
    source_headers: list[str],
    row: list[Any],
    row_index: int,
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "id": row_index + 1,
        "index": row_index + 1,
        "bankId": bank_id,
        "extraFields": {},
    }
    used_fixed_fields: set[str] = set()
    has_synthetic_id = headers and headers[0] == "ID" and "ID" not in source_headers

    for column_index, column_name in enumerate(headers):
        if column_name == "ID" and has_synthetic_id:
            value = row_index + 1
        else:
            source_index = column_index - 1 if has_synthetic_id else column_index
            value = row[source_index] if source_index < len(row) else ""

        field, _ = resolve_field(column_name, kind)
        if field is None or should_use_extra(field, used_fixed_fields):
            key = extra_field(bank_name, kind, column_name, column_index)
            item["extraFields"][key] = to_text(value)
            continue

        item[field] = convert_value(field, value)

    item.setdefault("isCheck", False)
    item.setdefault("incomeAttribute", "")
    item.setdefault("minMoney", None)
    item.setdefault("maxMoney", None)
    item.setdefault("floutLength", None)
    item.setdefault("startDay", None)
    item.setdefault("endDay", None)
    item.setdefault("tradeHoliday", False)
    item.setdefault("tradeWeekend", False)

    item["id"] = row_index + 1
    item["index"] = row_index + 1
    item["bankId"] = bank_id

    if kind == REFERENCE_KIND:
        item.setdefault("percentMonth", None)
    else:
        item.setdefault("fixDay", "")
        item.setdefault("reCnt", "")

    return item


def default_config() -> dict[str, Any]:
    return {
        "selectIndex": 0,
        "startTime": "2023-04-01T00:00:00",
        "endTime": "2026-01-01T00:00:00",
        "allInMoney": 30000.0,
        "allOutMoney": 20000.0,
        "lastMoney": 10000.0,
        "minInMoneyMonth1": 3000.0,
        "maxInMoneyMonth1": 3000.0,
        "minOutMoneyMonth1": 2000.0,
        "maxOutMoneyMonth1": 2000.0,
        "minInMoneyMonth2": 3000.0,
        "maxInMoneyMonth2": 3000.0,
        "minOutMoneyMonth2": 2000.0,
        "maxOutMoneyMonth2": 2000.0,
        "monthGenData": [],
    }


def main() -> None:
    column_doc = json.loads(COLUMN_OUTPUT.read_text(encoding="utf-8-sig"))
    seed_doc = json.loads(SEED_OUTPUT.read_text(encoding="utf-8-sig"))
    column_doc.setdefault("banks", {})
    seed_doc.setdefault("banks", {})
    seed_doc.setdefault("bankNameKeys", {}).update(PERSONAL_BANK_SEED_KEYS)

    total_files = 0
    total_reference_rows = 0
    total_const_rows = 0

    for source_dir, bank_ids, label in [
        (PERSONAL_SOURCE_DIR, PERSONAL_BANK_IDS, "个人银行"),
        (SOURCE_DIR, CORPORATE_BANK_IDS, "对公银行"),
    ]:
        extracted: dict[str, dict[str, Any]] = {}
        files = sorted(path for path in source_dir.glob("*.xlsx") if not path.name.startswith("~$"))
        total_files += len(files)

        for path in files:
            workbook_data = read_workbook(path)
            bank_name = workbook_data["bank"]
            if bank_name not in bank_ids:
                raise ValueError(f"未知{label}文件名: {path.name} -> {bank_name}")

            bank_data = extracted.setdefault(
                bank_name,
                {
                    "reference": None,
                    "const": None,
                },
            )
            if workbook_data["kind"] == CONST_KIND:
                bank_data["const"] = workbook_data
            else:
                bank_data["reference"] = workbook_data

        missing = [
            bank_name
            for bank_name in bank_ids
            if bank_name not in extracted
            or extracted[bank_name]["reference"] is None
            or extracted[bank_name]["const"] is None
        ]
        if missing:
            raise ValueError(f"缺少{label}参照或固定文件: " + ", ".join(missing))

        for bank_name, bank_id in bank_ids.items():
            reference = extracted[bank_name]["reference"]
            const = extracted[bank_name]["const"]
            assert reference is not None and const is not None

            column_doc["banks"][bank_name] = {
                "reference": reference["headers"],
                "const": const["headers"],
            }

            references = [
                create_rule(
                    bank_name,
                    bank_id,
                    REFERENCE_KIND,
                    reference["headers"],
                    reference["sourceHeaders"],
                    row,
                    row_index,
                )
                for row_index, row in enumerate(reference["rows"])
            ]
            const_items = [
                create_rule(
                    bank_name,
                    bank_id,
                    CONST_KIND,
                    const["headers"],
                    const["sourceHeaders"],
                    row,
                    row_index,
                )
                for row_index, row in enumerate(const["rows"])
            ]

            total_reference_rows += len(references)
            total_const_rows += len(const_items)

            seed_doc["bankNameKeys"][bank_name] = str(bank_id)
            seed_doc["banks"][str(bank_id)] = {
                "config": default_config(),
                "references": references,
                "constItems": const_items,
            }

    column_doc["source"] = (
        "D:/Projects/财务系统开发资料/参照明细/*.xlsx; "
        "D:/Projects/财务系统开发资料/参照明细/对公银行/*.xlsx"
    )
    seed_doc["source"] = column_doc["source"]

    COLUMN_OUTPUT.write_text(json.dumps(column_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    SEED_OUTPUT.write_text(json.dumps(seed_doc, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "files": total_files,
                "banks": len(PERSONAL_BANK_IDS) + len(CORPORATE_BANK_IDS),
                "referenceRows": total_reference_rows,
                "constRows": total_const_rows,
                "totalColumnBanks": len(column_doc["banks"]),
                "totalSeedBanks": len(seed_doc["banks"]),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
